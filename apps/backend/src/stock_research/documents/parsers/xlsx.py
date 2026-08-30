from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory

XLSX_PARSER_VERSION = "openpyxl:1.0.0"


class XlsxParser:
    """使用 openpyxl 解析 XLSX 工作表。"""

    version = XLSX_PARSER_VERSION

    def parse(self, data: bytes, filename: str) -> str:
        with TemporaryDirectory() as tmp:
            path = Path(tmp) / (Path(filename).name or "document.xlsx")
            path.write_bytes(data)
            return self._parse_path(path)

    def _parse_path(self, path: Path) -> str:
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = [
                "\t".join("" if value is None else str(value) for value in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            return "\n".join(rows)
        finally:
            workbook.close()
